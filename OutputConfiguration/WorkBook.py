from pathlib import Path
from time import sleep

import xlsxwriter as xlwr
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

fileName = "tablica.xlsx"
name_sheet1 = 'Общие характеристики'
name_sheet2 = 'Характеристики дисков'
allconfiguration = []
diskconfiguration = []
wb = ""
sheet1 = ""
sheet2 = ""
rowinputall = ""
rowinputdisk = ""

def create():
    fileObj = Path(fileName)
    raw = 0

    if fileObj.is_file() == False:
        workbook = xlwr.Workbook(fileName, {'strings_to_numbers': True})
        worksheet = workbook.add_worksheet(name_sheet1)
        worksheet2 = workbook.add_worksheet(name_sheet2)
        bold = workbook.add_format({'bold': True})
        while raw < len(allconfiguration):
            worksheet.write(0, raw, allconfiguration[raw][0], bold)
            raw += 1
        raw = 0
        while raw < 5:
            if raw < 4:
                worksheet2.write(0, raw, diskconfiguration[raw][0], bold)
            else:
                i = 0
                while i<len(diskconfiguration[4]):
                    worksheet2.write(0, raw+i, diskconfiguration[raw][i][0], bold)
                    i+=1
            raw += 1
        workbook.close()
def init():
    wb = load_workbook('./' + fileName)
    return wb, wb[name_sheet1], wb[name_sheet2]
    
def create_hyperlink(rowin, rowend):
    hyperlinkfrom = []
    hyperlinkto = []
    for i in range(21,28,2):
        if allconfiguration[i][1] == "":
            break
        hyperlinkfrom.append(i)
    for i in range(rowin, rowend+1):
        hyperlinkto.append(i)
    for index, hl in enumerate(hyperlinkfrom):
        sheet1.cell(row = rowinputall, column = hl+1).hyperlink = fileName +"#'" + name_sheet2 + "'!E" + str(hyperlinkto[index]) + ':L' + str(hyperlinkto[index])
        
    
def checklastrecord():
    rowinputall = 1
    rowinputdisk = 1
    checklastrecord = 'V'
    while checklastrecord == 'V':
        rowinputall += 1
        cell_obj = sheet1.cell(row=rowinputall, column=1)
        checklastrecord = cell_obj.value
    checklastrecord = 'V'
    while checklastrecord == 'V':
        rowinputdisk += 1
        cell_obj = sheet2.cell(row=rowinputdisk, column=1)
        checklastrecord = cell_obj.value
    return rowinputall, rowinputdisk
        
def configurationdisk():
    columninput = 0
    rowin = rowinputdisk
    i = 4
    
    while columninput < 4:
        columninput += 1
        sheet2.cell(row = rowin, column =columninput).value = diskconfiguration[columninput-1][1]
        
    while i < len(diskconfiguration):
        for inputinfo in diskconfiguration[i]:
            columninput += 1
            sheet2.cell(row = rowin, column =columninput).value = inputinfo[1]
        sheet2.cell(row = rowin, column =1).value = "V"
        sheet2.cell(row=rowin, column=columninput).border = Border(right=Side(border_style="medium", color="000000"))
        sheet2.cell(row = rowin, column =columninput+1).value = " "
        columninput = 4
        rowin += 1
        i += 1
        
    for i in range(1, 13):
        sheet2.cell(row=rowin-1, column=i).border = Border(bottom=Side(border_style="medium", color="000000"))
        if i == 12:
            sheet2.cell(row=rowin-1, column=i).border = Border(right=Side(border_style="medium", color="000000"), bottom=Side(border_style="medium", color="000000"))
    if rowinputdisk != rowin-1:
        for i in range(2, 5):
            sheet2.merge_cells(start_row=rowinputdisk, start_column=i, end_row=rowin-1, end_column=i)
            sheet2.cell(row=rowinputdisk, column = i).alignment = Alignment(horizontal="center", vertical="center")
    return rowinputdisk, rowin-1

def configuration():
    columninput = 0

    while columninput < len(allconfiguration):
        columninput += 1
        sheet1.cell(row = rowinputall, column =columninput).value = allconfiguration[columninput-1][1]
    sheet1.cell(row = rowinputall, column =columninput+1).value = " "
        
def close():
    while True:
        try:
            wb.save(filename = fileName)
            break
        except PermissionError:
            print('Файл XLSX открыт, либо уже обрабатывается. Повтор записи данных через 5 секунды')
            sleep(5)